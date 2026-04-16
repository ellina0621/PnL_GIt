from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# 唯一 Excel 檔案：使用者在這裡填資料，腳本也寫回這裡
MASTER_FILE = Path("trading-journal (2) - 每日損益重算.xlsx")
INITIAL_CAPITAL = 100000.0


# 興櫃股票歷史收盤補充（收盤價 sheet 沒有這些股票的早期記錄）
# 未來每日新收盤只要填 Col R，這裡不需要再更新
SUPPLEMENTAL_CLOSES: Dict[int, Dict[date, float]] = {
    3585: {
        date(2026, 4, 13): 30.1,
        date(2026, 4, 14): 49.2,
        date(2026, 4, 15): 62.0,
    },
    7828: {
        date(2026, 4, 1): 1310.0,
        date(2026, 4, 2): 1285.0,
        date(2026, 4, 7): 1325.0,
        date(2026, 4, 8): 1405.0,
        date(2026, 4, 9): 1415.0,
        date(2026, 4, 10): 1395.0,
        date(2026, 4, 13): 1325.0,
        date(2026, 4, 14): 1315.0,
        date(2026, 4, 15): 1325.0,
    },
}


@dataclass
class Trade:
    trade_id: int
    code: int
    name: str
    market: str
    industry: str
    buy_date: date
    sell_date: Optional[date]
    shares: float
    buy_px: float
    sell_px: Optional[float]
    buy_fee: float
    sell_fee_tax: float
    realized: Optional[float]
    note: Optional[str]
    today_close: Optional[float]  # Col R：每天填今日收盤（持倉用）

    @property
    def buy_cost(self) -> float:
        return self.shares * self.buy_px + self.buy_fee

    @property
    def net_sell_proceeds(self) -> Optional[float]:
        if self.realized is None:
            return None
        return self.buy_cost + self.realized


def as_date(value) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        raw = str(int(value))
        if len(raw) == 8:
            return date(int(raw[:4]), int(raw[4:6]), int(raw[6:8]))
    raise ValueError(f"Unsupported date value: {value!r}")


def as_code(value) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    return int(str(value).strip())


def as_float(value, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def load_trades(workbook) -> List[Trade]:
    """
    讀取交易記錄 sheet。
    部分欄位為 Excel 公式（trade_id = ROW()-2；realized = (賣-買)*股-費）。
    openpyxl save() 後公式快取清空，所以這裡自行計算，不依賴快取。
    """
    ws = workbook["交易記錄"]
    trades: List[Trade] = []
    for row in range(3, ws.max_row + 1):
        code = as_code(ws.cell(row, 2).value)
        if code is None:
            continue

        shares       = as_float(ws.cell(row, 9).value)
        buy_px       = as_float(ws.cell(row, 10).value)
        sell_px_raw  = ws.cell(row, 11).value
        sell_px      = None if sell_px_raw in (None, "") else float(sell_px_raw)
        buy_fee      = as_float(ws.cell(row, 12).value)
        sell_fee_tax = as_float(ws.cell(row, 13).value)

        # trade_id：公式 =IF(B="","",ROW()-2)；快取清空後自行計算
        tid_raw = ws.cell(row, 1).value
        if tid_raw in (None, ""):
            trade_id = row - 2
        else:
            trade_id = int(as_float(tid_raw))

        # realized：公式 =((K-J)*I)-L-M；快取清空且有賣出價時自行計算
        realized_raw = ws.cell(row, 14).value
        if realized_raw not in (None, ""):
            realized = float(realized_raw)
        elif sell_px is not None and buy_px > 0 and shares > 0:
            realized = (sell_px - buy_px) * shares - buy_fee - sell_fee_tax
        else:
            realized = None  # 尚未賣出

        # today_close：Col R（使用者每日手填，非公式）
        tc_raw = ws.cell(row, 18).value
        today_close = None if tc_raw in (None, "") else float(tc_raw)

        trades.append(Trade(
            trade_id=trade_id,
            code=code,
            name=str(ws.cell(row, 3).value or ""),
            market=str(ws.cell(row, 4).value or ""),
            industry=str(ws.cell(row, 5).value or ""),
            buy_date=as_date(ws.cell(row, 6).value),
            sell_date=as_date(ws.cell(row, 7).value),
            shares=shares,
            buy_px=buy_px,
            sell_px=sell_px,
            buy_fee=buy_fee,
            sell_fee_tax=sell_fee_tax,
            realized=realized,
            note=ws.cell(row, 17).value,
            today_close=today_close,
        ))
    return trades


def load_local_closes(workbook) -> Tuple[Dict[int, Dict[date, float]], List[date]]:
    ws = workbook["收盤價"]
    closes: Dict[int, Dict[date, float]] = {}
    trading_dates = set()
    for row in range(2, ws.max_row + 1):
        dt   = as_date(ws.cell(row, 1).value)
        code = as_code(ws.cell(row, 2).value)
        px   = ws.cell(row, 3).value
        if dt is None or code is None or px in (None, ""):
            continue
        trading_dates.add(dt)
        closes.setdefault(code, {})[dt] = float(px)
    return closes, sorted(trading_dates)


def merge_all_closes(
    local_closes: Dict[int, Dict[date, float]],
    trades: List[Trade],
    today: date,
) -> Tuple[Dict[int, Dict[date, float]], List[Tuple[date, int, float]]]:
    """
    合併三個來源：
    1. 收盤價 sheet（歷史）
    2. SUPPLEMENTAL_CLOSES（興櫃補充）
    3. 交易記錄 Col R（今日）
    回傳 merged closes 與 [(date, code, px)] 的歸檔清單。
    """
    merged = {code: series.copy() for code, series in local_closes.items()}
    for code, series in SUPPLEMENTAL_CLOSES.items():
        merged.setdefault(code, {}).update(series)
    to_archive: List[Tuple[date, int, float]] = []
    for trade in trades:
        if trade.today_close is not None and trade.sell_date is None:
            merged.setdefault(trade.code, {})[today] = trade.today_close
            to_archive.append((today, trade.code, trade.today_close))
    return merged, to_archive


def price_as_of(
    close_series: Dict[date, float], current_date: date
) -> Tuple[Optional[float], Optional[date]]:
    eligible = [dt for dt in close_series if dt <= current_date]
    if not eligible:
        return None, None
    latest = max(eligible)
    return close_series[latest], latest


def style_header(ws, header_row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 24)


def remove_sheet_if_exists(workbook, name: str) -> None:
    if name in workbook.sheetnames:
        del workbook[name]


def ensure_col_r_header(ws) -> None:
    """若交易記錄第 2 列 Col R 還沒有「今日收盤」就加上。"""
    if ws.cell(2, 18).value == "今日收盤":
        return
    cell = ws.cell(2, 18)
    cell.value = "今日收盤"
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    print("  已新增「今日收盤」欄位標題（Col R）")


def append_new_closes(ws_close, to_archive: List[Tuple[date, int, float]],
                      existing_closes: Dict[int, Dict[date, float]]) -> None:
    """把 Col R 的今日收盤追加到收盤價 sheet（避免重複）。"""
    added = 0
    for dt, code, px in to_archive:
        if code in existing_closes and dt in existing_closes[code]:
            continue
        ws_close.append([int(dt.strftime("%Y%m%d")), code, px])
        added += 1
    if added:
        print(f"  歸檔 {added} 筆今日收盤到收盤價 sheet")


def bake_formulas(snapshot: Dict[Tuple[int, int], object], ws_write) -> None:
    """
    把 ws_write 裡的公式格（以 = 開頭）替換成 snapshot 裡的靜態值。
    解決 openpyxl save() 清除公式快取的問題：存一次後全部變靜態值，
    下次讀取就不需要快取，腳本可以無限重跑。
    """
    baked = 0
    for row in ws_write.iter_rows():
        for write_cell in row:
            if isinstance(write_cell.value, str) and write_cell.value.startswith("="):
                static_val = snapshot.get((write_cell.row, write_cell.column))
                write_cell.value = static_val  # formula → static value (or None)
                baked += 1
    if baked:
        print(f"  Baked {baked} formula cells in {ws_write.title}")


def main() -> None:
    if not MASTER_FILE.exists():
        raise FileNotFoundError(MASTER_FILE)

    today = date.today()

    # ── 讀資料（data_only=True；公式值若快取清空則自行計算）──────────────────
    wb_data = load_workbook(MASTER_FILE, data_only=True)
    trades = load_trades(wb_data)
    local_closes, local_dates = load_local_closes(wb_data)
    # 把交易記錄的格值存成 dict，供之後 bake 用（close 後 worksheet 就失效）
    trade_cell_snapshot: Dict[Tuple[int, int], object] = {
        (cell.row, cell.column): cell.value
        for row in wb_data["交易記錄"].iter_rows()
        for cell in row
    }
    wb_data.close()

    closes, to_archive = merge_all_closes(local_closes, trades, today)

    # ── 計算日期範圍 ──────────────────────────────────────────────────────────
    all_sell_dates = [t.sell_date for t in trades if t.sell_date is not None]
    start_date = date(2025, 12, 24)
    end_candidates = [
        max(local_dates) if local_dates else start_date,
        max(t.buy_date for t in trades),
    ]
    if all_sell_dates:
        end_candidates.append(max(all_sell_dates))
    if to_archive:
        end_candidates.append(today)
    end_date = max(end_candidates)

    event_dates = {
        t.buy_date
        for t in trades
        if start_date <= t.buy_date <= end_date
    } | {
        t.sell_date
        for t in trades
        if t.sell_date is not None and start_date <= t.sell_date <= end_date
    }

    today_set: set[date] = set()
    if to_archive and start_date <= today <= end_date:
        today_set = {today}

    calendar = sorted(
        {dt for dt in local_dates if start_date <= dt <= end_date}
        | event_dates
        | today_set
    )

    missing_codes = sorted({t.code for t in trades if t.code not in closes})
    if missing_codes:
        raise ValueError(f"Missing close series for codes: {missing_codes}")

    # ── 開啟檔案準備寫入 ──────────────────────────────────────────────────────
    wb = load_workbook(MASTER_FILE)
    # 把交易記錄的公式格全部烘焙成靜態值（避免 openpyxl save 清除快取）
    bake_formulas(trade_cell_snapshot, wb["交易記錄"])
    ensure_col_r_header(wb["交易記錄"])
    append_new_closes(wb["收盤價"], to_archive, local_closes)

    remove_sheet_if_exists(wb, "每日損益重算")
    remove_sheet_if_exists(wb, "每日明細重算")
    summary_ws = wb.create_sheet("每日損益重算")
    detail_ws  = wb.create_sheet("每日明細重算")

    # ── Summary sheet ─────────────────────────────────────────────────────────
    summary_ws["A1"] = "每日損益重算"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = "計算口徑：期末權益 = 現金餘額 + 持倉市值；未實現損益含買進手續費。"
    summary_ws["A3"] = (
        "收盤價來源：收盤價 sheet（歷史）；Col R 今日收盤（當日持倉）；缺價則沿用最近收盤。"
    )
    summary_ws.append([])
    summary_ws.append([
        "日期", "當日已實現損益", "累積已實現損益", "期末未實現損益",
        "總損益", "現金餘額", "持倉市值", "權益總值", "日權益變動",
        "持倉檔數", "已收盤價覆蓋說明",
    ])
    style_header(summary_ws, 5)

    # ── Detail sheet ──────────────────────────────────────────────────────────
    detail_ws["A1"] = "每日明細重算"
    detail_ws["A1"].font = Font(bold=True, size=14)
    detail_ws["A2"] = "同日賣出者，當天列入已實現損益，期末未實現損益為 0。"
    detail_ws.append([])
    detail_ws.append([
        "日期", "交易ID", "股票代號", "股票名稱", "市場別", "產業別",
        "事件", "股數", "買進日期", "賣出日期", "買進均價",
        "當日收盤價", "收盤價日期", "當日持倉市值",
        "當日已實現損益", "期末未實現損益", "當日說明",
    ])
    style_header(detail_ws, 4)

    # ── 主計算迴圈 ───────────────────────────────────────────────────────────
    cash = INITIAL_CAPITAL
    cumulative_realized = 0.0
    detail_rows: List[List] = []

    buy_events:  Dict[date, List[Trade]] = {}
    sell_events: Dict[date, List[Trade]] = {}
    for trade in trades:
        buy_events.setdefault(trade.buy_date, []).append(trade)
        if trade.sell_date:
            sell_events.setdefault(trade.sell_date, []).append(trade)

    for trade in trades:
        if trade.buy_date < start_date:
            cash -= trade.buy_cost
        if trade.sell_date and trade.sell_date < start_date:
            proceeds = trade.net_sell_proceeds
            if proceeds is None:
                raise ValueError(f"Trade {trade.trade_id} missing realized pnl")
            cash += proceeds
            cumulative_realized += trade.realized or 0.0

    prev_equity: Optional[float] = None

    for current_date in calendar:
        for trade in buy_events.get(current_date, []):
            cash -= trade.buy_cost

        day_realized = 0.0
        for trade in sell_events.get(current_date, []):
            proceeds = trade.net_sell_proceeds
            if proceeds is None:
                raise ValueError(f"Trade {trade.trade_id} missing realized pnl")
            cash += proceeds
            day_realized        += trade.realized or 0.0
            cumulative_realized += trade.realized or 0.0

        open_positions = [
            t for t in trades
            if t.buy_date <= current_date
            and (t.sell_date is None or current_date < t.sell_date)
        ]

        position_value   = 0.0
        unrealized_total = 0.0
        carry_notes: List[str] = []

        for trade in open_positions:
            series = closes[trade.code]
            close_px, close_dt = price_as_of(series, current_date)
            if close_px is None or close_dt is None:
                raise ValueError(
                    f"No price for {trade.code} on or before {current_date.isoformat()}"
                )
            market_value     = trade.shares * close_px
            unrealized       = market_value - trade.buy_cost
            position_value  += market_value
            unrealized_total += unrealized
            if close_dt != current_date:
                carry_notes.append(f"{trade.code}沿用{close_dt.isoformat()}收盤")
            detail_rows.append([
                current_date, trade.trade_id, trade.code, trade.name,
                trade.market, trade.industry, "持有中", trade.shares,
                trade.buy_date, trade.sell_date, trade.buy_px,
                close_px, close_dt, market_value, 0.0, unrealized, "當日期末持有部位",
            ])

        for trade in sell_events.get(current_date, []):
            detail_rows.append([
                current_date, trade.trade_id, trade.code, trade.name,
                trade.market, trade.industry, "當日賣出", trade.shares,
                trade.buy_date, trade.sell_date, trade.buy_px,
                trade.sell_px, current_date, 0.0, trade.realized or 0.0, 0.0,
                "已由未實現轉入已實現",
            ])

        total_pnl  = cumulative_realized + unrealized_total
        equity     = cash + position_value
        day_change = None if prev_equity is None else equity - prev_equity

        diff = round(total_pnl - (equity - INITIAL_CAPITAL), 6)
        if abs(diff) > 0.01:
            raise ValueError(
                f"PnL mismatch on {current_date.isoformat()}: "
                f"pnl={total_pnl:.2f}, equity_delta={equity - INITIAL_CAPITAL:.2f}"
            )
        prev_equity = equity

        coverage_note = "；".join(sorted(set(carry_notes))) if carry_notes else ""
        summary_ws.append([
            current_date, day_realized, cumulative_realized, unrealized_total,
            total_pnl, cash, position_value, equity, day_change,
            len(open_positions), coverage_note,
        ])

    for row in detail_rows:
        detail_ws.append(row)

    # ── 格式化 ───────────────────────────────────────────────────────────────
    date_fmt  = "yyyy-mm-dd"
    money_fmt = "#,##0.00;(#,##0.00)"
    int_fmt   = "#,##0"

    for row in summary_ws.iter_rows(min_row=6, max_row=summary_ws.max_row):
        row[0].number_format = date_fmt
        for cell in row[1:9]:
            cell.number_format = money_fmt
        row[9].number_format = int_fmt

    for row in detail_ws.iter_rows(min_row=5, max_row=detail_ws.max_row):
        row[0].number_format = date_fmt
        row[8].number_format = date_fmt
        row[9].number_format = date_fmt
        row[12].number_format = date_fmt
        row[7].number_format  = "#,##0.####"
        for idx in [10, 11, 13, 14, 15]:
            row[idx].number_format = money_fmt

    summary_ws.freeze_panes = "A6"
    detail_ws.freeze_panes  = "A5"
    autofit_columns(summary_ws)
    autofit_columns(detail_ws)

    wb.save(MASTER_FILE)
    print(f"Wrote {MASTER_FILE}")
    print(f"Summary rows: {summary_ws.max_row - 5}")
    print(f"Detail rows:  {detail_ws.max_row - 4}")


if __name__ == "__main__":
    main()
